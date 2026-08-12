"""
Exports professionnels du module Stock : Excel (openpyxl) et PDF (reportlab).
Helpers génériques (tableaux) + génération PDF d'un document de vente.
"""
from io import BytesIO

from django.http import HttpResponse


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def excel_response(filename, titre, colonnes, lignes, total_row=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = titre[:31]

    entete_fill = PatternFill(start_color='198754', end_color='198754', fill_type='solid')
    gras_blanc = Font(bold=True, color='FFFFFF')

    ws.append([titre])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(colonnes)))
    ws['A1'].font = Font(bold=True, size=14)

    ws.append([])
    ws.append(list(colonnes))
    for cell in ws[3]:
        cell.font = gras_blanc
        cell.fill = entete_fill
        cell.alignment = Alignment(horizontal='center')

    for ligne in lignes:
        ws.append(list(ligne))

    if total_row:
        ws.append(list(total_row))
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

    # Largeur des colonnes approximative
    for i, _ in enumerate(colonnes, start=1):
        from openpyxl.utils import get_column_letter
        longueurs = [len(str(colonnes[i - 1]))]
        for ligne in lignes:
            if i - 1 < len(ligne):
                longueurs.append(len(str(ligne[i - 1])))
        ws.column_dimensions[get_column_letter(i)].width = min(max(longueurs) + 3, 45)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------
def _styles():
    from reportlab.lib.styles import getSampleStyleSheet
    return getSampleStyleSheet()


def pdf_table_response(filename, titre, colonnes, lignes, entreprise=None, total_row=None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            topMargin=1 * cm, bottomMargin=1 * cm, leftMargin=1 * cm, rightMargin=1 * cm)
    styles = _styles()
    elements = []
    if entreprise is not None:
        elements.append(Paragraph(str(getattr(entreprise, 'nom_entreprise', '')), styles['Heading2']))
    elements.append(Paragraph(titre, styles['Heading1']))
    elements.append(Spacer(1, 0.4 * cm))

    data = [list(colonnes)] + [list(l) for l in lignes]
    if total_row:
        data.append(list(total_row))

    table = Table(data, repeatRows=1)
    style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#198754')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f2f7f4')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]
    if total_row:
        style.append(('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'))
        style.append(('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e0e0e0')))
    table.setStyle(TableStyle(style))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}.pdf"'
    return response


def pdf_document_vente(doc_vente):
    """Génère le PDF d'un devis / proforma / facture (avec lignes et totaux)."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    entreprise = doc_vente.entreprise
    client = doc_vente.client
    buffer = BytesIO()
    pdf = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.2 * cm, bottomMargin=1.2 * cm,
                            leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    styles = _styles()
    el = []

    el.append(Paragraph(str(getattr(entreprise, 'nom_entreprise', '')), styles['Heading2']))
    el.append(Paragraph(f"{doc_vente.get_type_document_display().upper()} N° {doc_vente.reference}", styles['Heading1']))
    el.append(Paragraph(f"Date : {doc_vente.date_document:%d/%m/%Y}", styles['Normal']))
    el.append(Spacer(1, 0.3 * cm))
    el.append(Paragraph(f"<b>Client :</b> {client.nom}", styles['Normal']))
    if client.adresse:
        el.append(Paragraph(client.adresse, styles['Normal']))
    if client.telephone:
        el.append(Paragraph(f"Tél : {client.telephone}", styles['Normal']))
    el.append(Spacer(1, 0.5 * cm))

    data = [['Article', 'Qté', 'P.U. (GNF)', 'Remise', 'Montant (GNF)']]
    for l in doc_vente.lignes.all():
        data.append([
            l.article.designation, f"{l.quantite:.0f}", f"{l.prix_unitaire:,.0f}",
            f"{l.remise_pct:.0f} %", f"{l.montant:,.0f}",
        ])
    table = Table(data, colWidths=[8 * cm, 2 * cm, 3 * cm, 2 * cm, 3 * cm], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#198754')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
    ]))
    el.append(table)
    el.append(Spacer(1, 0.4 * cm))

    totaux = [
        ['Total HT', f"{doc_vente.total_ht:,.0f} GNF"],
        [f"TVA ({doc_vente.taux_tva:.0f} %)", f"{doc_vente.montant_tva:,.0f} GNF"],
        ['Total TTC', f"{doc_vente.montant_ttc:,.0f} GNF"],
    ]
    if doc_vente.est_facture:
        totaux.append(['Encaissé', f"{doc_vente.montant_paye:,.0f} GNF"])
        totaux.append(['Solde dû', f"{doc_vente.montant_du:,.0f} GNF"])
    t2 = Table(totaux, colWidths=[5 * cm, 4 * cm], hAlign='RIGHT')
    t2.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('LINEABOVE', (0, 2), (-1, 2), 0.6, colors.black),
        ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
    ]))
    el.append(t2)
    if doc_vente.notes:
        el.append(Spacer(1, 0.5 * cm))
        el.append(Paragraph(f"<i>{doc_vente.notes}</i>", styles['Normal']))

    pdf.build(el)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{doc_vente.reference}.pdf"'
    return response
