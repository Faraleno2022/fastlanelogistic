"""Utilitaires communs pour exports Excel/PDF et génération de fiches vierges.

Deux briques :
  - ``build_excel(title, columns, rows)`` -> HttpResponse XLSX
  - ``build_pdf(title, columns, rows, orientation)`` -> HttpResponse PDF

Chaque app (operations, facturation, rh, flotte) importe ces fonctions et
y pousse ses propres colonnes.
"""
from __future__ import annotations
import io
import os
from datetime import datetime, date
from decimal import Decimal
from typing import Callable, Iterable, Sequence

from django.conf import settings
from django.http import HttpResponse
from django.utils.text import slugify

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak,
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from .models import Parametres


# ---------------------------------------------------------------------------
# Identité de l'entreprise (logo + nom)
# ---------------------------------------------------------------------------

LOGO_CANDIDATE_PATHS = [
    os.path.join(settings.BASE_DIR, "static", "images", "logo_fastlane.png"),
    os.path.join(settings.BASE_DIR, "static", "images", "logo.png"),
    os.path.join(settings.BASE_DIR, "media", "logo_fastlane.png"),
]


def find_logo_path() -> str | None:
    for p in LOGO_CANDIDATE_PATHS:
        if os.path.exists(p):
            return p
    return None


# Cache mémoire pour le logo en filigrane (ImageReader avec alpha réduit).
_WATERMARK_CACHE: dict = {}


def _get_watermark_image() -> "ImageReader | None":
    """Retourne une version translucide du logo pour usage en filigrane.

    Le fichier est lu une seule fois puis mis en cache (on retourne un
    ImageReader prêt à passer à ``canvas.drawImage``). Si Pillow n'est
    pas disponible ou si aucun logo n'est trouvé, retourne ``None``.
    """
    if "img" in _WATERMARK_CACHE:
        return _WATERMARK_CACHE["img"]
    path = find_logo_path()
    if not path:
        _WATERMARK_CACHE["img"] = None
        return None
    try:
        from PIL import Image as PILImage
        img = PILImage.open(path).convert("RGBA")
        # Réduit l'opacité à ~8% pour un filigrane discret.
        alpha = img.split()[-1].point(lambda p: int(p * 0.08))
        img.putalpha(alpha)
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        reader = ImageReader(buf)
        _WATERMARK_CACHE["img"] = reader
        return reader
    except Exception:
        _WATERMARK_CACHE["img"] = None
        return None


def draw_watermark(canvas, pagesize, size_mm: float = 130) -> None:
    """Dessine le logo en filigrane centré sur la page (léger, rotation 25°)."""
    wm = _get_watermark_image()
    if not wm:
        return
    page_w, page_h = pagesize
    size = size_mm * mm
    canvas.saveState()
    try:
        canvas.translate(page_w / 2, page_h / 2)
        canvas.rotate(25)
        canvas.drawImage(
            wm, -size / 2, -size / 2, width=size, height=size,
            mask="auto", preserveAspectRatio=True,
        )
    finally:
        canvas.restoreState()


def get_company_info() -> dict:
    try:
        params = Parametres.load()
        return {
            "nom": params.societe_nom,
            "activite": params.activite,
            "devise": params.devise,
            "logo": find_logo_path(),
        }
    except Exception:
        return {
            "nom": getattr(settings, "SOCIETE_NOM", "Fastlane Logistic"),
            "activite": "Transport de Bauxite",
            "devise": getattr(settings, "SOCIETE_DEVISE", "GNF"),
            "logo": find_logo_path(),
        }


# ---------------------------------------------------------------------------
# Helpers de formatage
# ---------------------------------------------------------------------------

def fmt_value(v):
    """Rend une valeur directement exploitable par Excel et PDF."""
    if v is None:
        return ""
    if isinstance(v, (date, datetime)):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, Decimal):
        return float(v)
    return v


def fmt_cell_pdf(v) -> str:
    """Rend une chaîne imprimable pour le PDF."""
    if v is None:
        return ""
    if isinstance(v, (date, datetime)):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, Decimal):
        # nombres >= 1000 : séparateur d'espaces
        as_float = float(v)
        if as_float == int(as_float):
            return f"{int(as_float):,}".replace(",", " ")
        return f"{as_float:,.2f}".replace(",", " ")
    if isinstance(v, (int, float)):
        if float(v) == int(v):
            return f"{int(v):,}".replace(",", " ")
        return f"{float(v):,.2f}".replace(",", " ")
    return str(v)


# ---------------------------------------------------------------------------
# Export Excel
# ---------------------------------------------------------------------------

def build_excel(
    title: str,
    columns: Sequence[str],
    rows: Iterable[Sequence],
    *,
    filename: str | None = None,
    sheet_name: str = "Données",
    sous_titre: str | None = None,
    totaux: dict | None = None,
) -> HttpResponse:
    """Génère un fichier XLSX avec en-tête Fastlane Logistic.

    - ``columns`` : liste de libellés de colonnes
    - ``rows``    : iterable de tuples/listes de valeurs
    - ``totaux``  : {index_colonne: libelle_ou_valeur} ajouté en pied de tableau
    """
    info = get_company_info()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:30]

    # Styles
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=16, color="1F4E78")
    subtitle_font = Font(italic=True, size=10, color="555555")
    total_fill = PatternFill("solid", fgColor="E7E6E6")
    total_font = Font(bold=True)
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    n_cols = len(columns)
    last_col = get_column_letter(n_cols)

    # Logo (facultatif)
    row_title = 1
    if info["logo"]:
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(info["logo"])
            img.width, img.height = 120, 60
            ws.add_image(img, "A1")
            ws.row_dimensions[1].height = 50
            ws.row_dimensions[2].height = 20
            row_title = 2
        except Exception:
            pass

    # En-tête société + titre
    ws.cell(row=row_title, column=2, value=info["nom"]).font = title_font
    ws.merge_cells(start_row=row_title, start_column=2, end_row=row_title, end_column=n_cols)
    ws.cell(row=row_title + 1, column=2, value=f"{info['activite']} — {title}").font = subtitle_font
    ws.merge_cells(start_row=row_title + 1, start_column=2, end_row=row_title + 1, end_column=n_cols)
    if sous_titre:
        ws.cell(row=row_title + 2, column=2, value=sous_titre).font = subtitle_font
        ws.merge_cells(start_row=row_title + 2, start_column=2, end_row=row_title + 2, end_column=n_cols)
        row_title += 1
    ws.cell(row=row_title + 2, column=2,
            value=f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = subtitle_font

    header_row = row_title + 4

    # En-tête de colonnes
    for i, label in enumerate(columns, start=1):
        c = ws.cell(row=header_row, column=i, value=label)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    ws.row_dimensions[header_row].height = 24

    # Lignes de données
    r = header_row + 1
    col_widths = [max(12, len(str(lbl)) + 2) for lbl in columns]
    for row in rows:
        for i, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=i, value=fmt_value(v))
            cell.border = border
            if isinstance(v, (int, float, Decimal)):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            text = str(fmt_value(v))
            col_widths[i - 1] = max(col_widths[i - 1], min(40, len(text) + 2))
        r += 1

    # Pied de totaux
    if totaux:
        for i in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=i, value=totaux.get(i, ""))
            cell.fill = total_fill
            cell.font = total_font
            cell.border = border
            if isinstance(totaux.get(i), (int, float, Decimal)):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")

    # Largeur des colonnes
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Figer la ligne d'en-tête
    ws.freeze_panes = f"A{header_row + 1}"

    # Mise en page impression
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.print_options.horizontalCentered = True

    # Réponse HTTP
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = (filename or slugify(title) or "export") + ".xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    return response


# ---------------------------------------------------------------------------
# Export PDF
# ---------------------------------------------------------------------------

def _pdf_header_flowables(title: str, sous_titre: str | None = None) -> list:
    """Bloc d'en-tête (logo + nom société + titre) réutilisable."""
    info = get_company_info()
    styles = getSampleStyleSheet()

    style_nom = ParagraphStyle(
        "CompanyName", parent=styles["Title"], fontSize=16, leading=18,
        textColor=colors.HexColor("#1F4E78"), alignment=TA_LEFT, spaceAfter=0,
    )
    style_sub = ParagraphStyle(
        "Sub", parent=styles["Normal"], fontSize=9,
        textColor=colors.HexColor("#555555"), alignment=TA_LEFT,
    )
    style_title = ParagraphStyle(
        "Titre", parent=styles["Heading2"], fontSize=13,
        textColor=colors.HexColor("#1F4E78"), alignment=TA_CENTER, spaceBefore=4, spaceAfter=4,
    )

    # Ligne haut : [ logo | société + activité ] puis titre centré dessous
    left_cell = ""
    if info["logo"]:
        try:
            left_cell = Image(info["logo"], width=35 * mm, height=17 * mm)
        except Exception:
            left_cell = ""

    right_cell = [
        Paragraph(info["nom"], style_nom),
        Paragraph(f"{info['activite']}", style_sub),
        Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}", style_sub),
    ]

    header_tbl = Table(
        [[left_cell, right_cell]],
        colWidths=[40 * mm, None],
    )
    header_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))

    flow = [header_tbl, Spacer(1, 4 * mm),
            Paragraph(title, style_title)]
    if sous_titre:
        flow.append(Paragraph(sous_titre, style_sub))
    flow.append(Spacer(1, 3 * mm))
    return flow


def build_pdf(
    title: str,
    columns: Sequence[str],
    rows: Iterable[Sequence],
    *,
    filename: str | None = None,
    orientation: str = "landscape",
    sous_titre: str | None = None,
    totaux: list | None = None,
    col_widths: Sequence[float] | None = None,
) -> HttpResponse:
    """Génère un PDF tabulaire avec en-tête société.

    - ``totaux`` : liste de cellules formant la dernière ligne (mise en gras)
    - ``col_widths`` : si fourni, liste en mm pour chaque colonne
    """
    pagesize = landscape(A4) if orientation == "landscape" else A4
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=pagesize,
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=10 * mm, bottomMargin=10 * mm,
        title=title,
    )

    flow = _pdf_header_flowables(title, sous_titre)

    # Préparation des données
    data = [list(columns)]
    for row in rows:
        data.append([fmt_cell_pdf(v) for v in row])
    if totaux:
        data.append([fmt_cell_pdf(v) for v in totaux])

    widths = None
    if col_widths:
        widths = [w * mm for w in col_widths]
    tbl = Table(data, colWidths=widths, repeatRows=1)

    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#BFBFBF")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2 if totaux else -1),
         [colors.white, colors.HexColor("#F7F9FB")]),
    ])
    if totaux:
        style.add("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E7E6E6"))
        style.add("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold")
    tbl.setStyle(style)

    flow.append(tbl)

    # Pied de page avec pagination
    def _on_page(canvas, d):
        draw_watermark(canvas, pagesize)
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#888888"))
        footer = f"{get_company_info()['nom']} — Page {canvas.getPageNumber()}"
        canvas.drawRightString(pagesize[0] - 10 * mm, 6 * mm, footer)
        canvas.restoreState()

    doc.build(flow, onFirstPage=_on_page, onLaterPages=_on_page)

    fname = (filename or slugify(title) or "export") + ".pdf"
    response = HttpResponse(buf.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    return response


# ---------------------------------------------------------------------------
# Fiches vierges terrain (pré-imprimées pour agents)
# ---------------------------------------------------------------------------

def build_blank_fiches(
    fiche_type: str,
    nb_copies: int = 1,
    nb_lignes: int = 12,
) -> HttpResponse:
    """Génère un PDF de fiches vierges à remplir à la main.

    ``fiche_type`` ∈ {"bon_transport", "voyage_bauxite", "carburant", "panne"}

    Mise en page : tableau multi-lignes en A4 paysage. Chaque colonne = un champ,
    chaque ligne = une saisie (ex: une prise de carburant). ``nb_copies`` est
    le nombre de pages à générer, ``nb_lignes`` le nombre de lignes vides
    par page.
    """
    buf = io.BytesIO()
    pagesize = landscape(A4)
    doc = SimpleDocTemplate(
        buf, pagesize=pagesize,
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=10 * mm, bottomMargin=10 * mm,
        title=f"Fiche vierge — {fiche_type}",
    )
    flow = []
    definitions = {
        "bon_transport": {
            "titre": "BON DE TRANSPORT",
            "champs": [
                ("N° Bon", 8), ("Date", 6),
                ("Prénom chauffeur", 8), ("Nom chauffeur", 8),
                ("Téléphone", 8),
                ("Plaque camion", 8), ("Carte d'entrée", 8),
                ("Lieu de chargement", 12),
                ("Heure départ", 6), ("Pesée début", 6), ("Pesée fin", 6),
                ("Quantité (T)", 6),
                ("Client / Projet", 10),
                ("Observations", 15),
            ],
            "signatures": ["Signature chauffeur", "Signature agent de terrain"],
        },
        "voyage_bauxite": {
            "titre": "VOYAGE BAUXITE — FICHE DE SUIVI",
            "champs": [
                ("Date", 6),
                ("Camion (code)", 8), ("Chauffeur", 10),
                ("Client", 8), ("Contrat", 8),
                ("Trajet (départ → arrivée)", 15),
                ("Distance (km)", 6), ("Tonnage (T)", 6),
                ("Tarif (GNF/T)", 8),
                ("N° Bon de livraison", 8),
                ("Observations", 15),
            ],
            "signatures": ["Signature chauffeur", "Signature chef de chantier"],
        },
        "carburant": {
            "titre": "PRISE DE CARBURANT",
            "champs": [
                ("Date", 6), ("Heure", 4),
                ("Camion (code)", 8), ("Chauffeur", 10),
                ("Km au tableau de bord", 8), ("Km avant prise", 8),
                ("Nb L avant prise", 8), ("Nb L après prise", 8),
                ("Prix unitaire (GNF/L)", 8),
                ("Station / Lieu", 10),
                ("Contrat / Projet", 10),
                ("Observations", 15),
            ],
            "signatures": ["Signature chauffeur", "Signature pompiste"],
        },
        "panne": {
            "titre": "PANNE / RÉPARATION",
            "champs": [
                ("Date", 6),
                ("Camion (code)", 8),
                ("Type de panne", 10),
                ("Pièce remplacée", 12),
                ("Fournisseur / Garage", 12),
                ("Coût pièces (GNF)", 8),
                ("Coût main d'œuvre (GNF)", 8),
                ("Durée immobilisation (jours)", 6),
                ("Contrat / Projet", 10),
                ("Observations", 15),
            ],
            "signatures": ["Signature mécanicien", "Signature responsable"],
        },
    }
    spec = definitions.get(fiche_type)
    if not spec:
        raise ValueError(f"Type de fiche inconnu: {fiche_type}")

    styles = getSampleStyleSheet()
    style_titre = ParagraphStyle(
        "T", parent=styles["Heading1"], fontSize=14, alignment=TA_CENTER,
        textColor=colors.HexColor("#1F4E78"), spaceAfter=3,
    )
    style_header = ParagraphStyle(
        "TH", parent=styles["Normal"], fontSize=7.5, leading=9,
        textColor=colors.white, alignment=TA_CENTER, fontName="Helvetica-Bold",
    )
    style_label = ParagraphStyle(
        "L", parent=styles["Normal"], fontSize=9,
        textColor=colors.HexColor("#333333"),
    )

    # Largeur utile du corps (A4 paysage, marges 10mm * 2) = 277 mm.
    body_width = pagesize[0] - 20 * mm
    # Colonne "#" fixe pour numéroter les lignes.
    col_num_width = 8 * mm
    # Répartition des colonnes au prorata des poids fournis dans `champs`.
    champs = spec["champs"]
    total_poids = sum(p for _, p in champs) or 1
    remaining_width = body_width - col_num_width
    col_widths = [col_num_width] + [
        (p / total_poids) * remaining_width for _, p in champs
    ]

    for copy_index in range(nb_copies):
        flow.extend(_pdf_header_flowables(spec["titre"]))

        # Construction du tableau :
        #   - Ligne 0 : en-têtes (bandeau bleu, texte blanc)
        #   - N lignes vides à remplir à la main
        header_row = [Paragraph("<b>#</b>", style_header)] + [
            Paragraph(f"<b>{label}</b>", style_header) for label, _ in champs
        ]
        data = [header_row]
        for i in range(1, nb_lignes + 1):
            data.append([str(i)] + [""] * len(champs))

        header_height = 9 * mm
        row_height = 8.5 * mm
        row_heights = [header_height] + [row_height] * nb_lignes

        tbl = Table(
            data,
            colWidths=col_widths,
            rowHeights=row_heights,
            repeatRows=1,
        )
        tbl.setStyle(TableStyle([
            # En-tête
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
            # Numéros de ligne (colonne 0 hors en-tête)
            ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#F2F6FA")),
            ("ALIGN", (0, 1), (0, -1), "CENTER"),
            ("VALIGN", (0, 1), (0, -1), "MIDDLE"),
            ("FONTSIZE", (0, 1), (0, -1), 8),
            ("TEXTCOLOR", (0, 1), (0, -1), colors.HexColor("#666666")),
            # Grille générale
            ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#1F4E78")),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#BFBFBF")),
            # Alternance lignes
            ("ROWBACKGROUNDS", (1, 1), (-1, -1),
             [colors.white, colors.HexColor("#FAFBFD")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, 0), 3),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 3),
        ]))

        flow.append(tbl)
        flow.append(Spacer(1, 3 * mm))

        # Bloc signatures : deux colonnes de largeur égale
        half = body_width / 2
        sig_data = [[Paragraph(f"<b>{s}</b>", style_label) for s in spec["signatures"]],
                    ["", ""]]
        sig_tbl = Table(sig_data, colWidths=[half, half])
        sig_tbl.setStyle(TableStyle([
            ("LINEBELOW", (0, 1), (-1, 1), 0.6, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 1), (-1, 1), 10),
            ("BOTTOMPADDING", (0, 1), (-1, 1), 0),
        ]))
        flow.append(sig_tbl)

        if copy_index < nb_copies - 1:
            flow.append(PageBreak())

    def _on_page(canvas, d):
        draw_watermark(canvas, pagesize)
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#888888"))
        footer = f"{get_company_info()['nom']} — Fiche terrain"
        canvas.drawRightString(pagesize[0] - 10 * mm, 6 * mm, footer)
        canvas.restoreState()

    doc.build(flow, onFirstPage=_on_page, onLaterPages=_on_page)

    response = HttpResponse(buf.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="fiche_{fiche_type}.pdf"'
    )
    return response


# ---------------------------------------------------------------------------
# Bon de transport individuel (imprimable — à remettre au chauffeur)
# ---------------------------------------------------------------------------

def build_bon_transport_pdf(bon, inline: bool = False) -> HttpResponse:
    """Génère un PDF professionnel pour un bon de transport individuel.

    Mise en page calquée sur le modèle officiel FASTLANE LOGISTIC :
      - En-tête 5 colonnes : société | N° BON | valeur | DATE | valeur
      - Bandeau titre "BON DE TRANSPORT BAUXITE — FASTLANE LOGISTIC"
      - Bloc principal sur 2 colonnes :
          * gauche (étroite) : sous-sections CHAUFFEUR / VÉHICULE / CLIENT-PROJET
          * droite (large)   : PESÉES ET QUANTITÉ + OBSERVATIONS + SIGNATURES
      - Avertissement légal en bas
      - Logo en filigrane sur toute la page

    ``bon``   : instance ``apps.operations.models.BonTransport``.
    ``inline``: si True, le PDF s'affiche dans le navigateur ; sinon il est
                téléchargé.
    """
    info = get_company_info()
    buf = io.BytesIO()
    pagesize = A4
    page_w, page_h = pagesize
    lm = rm = 12 * mm
    tm = bm = 12 * mm
    body_w = page_w - lm - rm

    doc = SimpleDocTemplate(
        buf, pagesize=pagesize,
        leftMargin=lm, rightMargin=rm,
        topMargin=tm, bottomMargin=bm,
        title=f"Bon de transport {bon.num_bon}",
    )

    styles = getSampleStyleSheet()
    NAVY = colors.HexColor("#1F4E78")
    LIGHT = colors.HexColor("#F2F6FA")
    GREY_BORDER = colors.HexColor("#BFBFBF")
    INK = colors.HexColor("#1F2937")
    MUTED = colors.HexColor("#666666")

    st_company = ParagraphStyle(
        "BtCompany", parent=styles["Normal"], fontSize=13, leading=15,
        textColor=NAVY, fontName="Helvetica-Bold", spaceAfter=0,
    )
    st_company_sub = ParagraphStyle(
        "BtCompanySub", parent=styles["Normal"], fontSize=8, leading=10,
        textColor=MUTED, fontName="Helvetica",
    )
    st_hdr_label = ParagraphStyle(
        "BtHdrLabel", parent=styles["Normal"], fontSize=7.5, leading=9,
        textColor=colors.white, fontName="Helvetica-Bold", alignment=TA_CENTER,
    )
    st_hdr_val = ParagraphStyle(
        "BtHdrVal", parent=styles["Normal"], fontSize=12, leading=14,
        textColor=NAVY, fontName="Helvetica-Bold", alignment=TA_CENTER,
    )
    st_title_band = ParagraphStyle(
        "BtTitleBand", parent=styles["Normal"], fontSize=13, leading=16,
        textColor=colors.white, fontName="Helvetica-Bold", alignment=TA_CENTER,
    )
    st_section = ParagraphStyle(
        "BtSection", parent=styles["Normal"], fontSize=9.5, leading=12,
        textColor=colors.white, fontName="Helvetica-Bold", alignment=TA_LEFT,
    )
    st_label = ParagraphStyle(
        "BtLabel", parent=styles["Normal"], fontSize=8, leading=10,
        textColor=MUTED, fontName="Helvetica-Bold",
    )
    st_val = ParagraphStyle(
        "BtVal", parent=styles["Normal"], fontSize=10, leading=12,
        textColor=INK, fontName="Helvetica-Bold",
    )
    st_weigh_head = ParagraphStyle(
        "BtWeighHead", parent=styles["Normal"], fontSize=8, leading=10,
        textColor=colors.white, fontName="Helvetica-Bold", alignment=TA_CENTER,
    )
    st_weigh_val = ParagraphStyle(
        "BtWeighVal", parent=styles["Normal"], fontSize=13, leading=15,
        textColor=INK, fontName="Helvetica-Bold", alignment=TA_CENTER,
    )
    st_sig_role = ParagraphStyle(
        "BtSigRole", parent=styles["Normal"], fontSize=8.5, leading=10,
        textColor=colors.white, fontName="Helvetica-Bold", alignment=TA_CENTER,
    )
    st_sig_note = ParagraphStyle(
        "BtSigNote", parent=styles["Normal"], fontSize=7.5, leading=9,
        textColor=MUTED, fontName="Helvetica-Oblique", alignment=TA_LEFT,
    )
    st_warn = ParagraphStyle(
        "BtWarn", parent=styles["Normal"], fontSize=8, leading=10,
        textColor=colors.HexColor("#7A1F1F"), fontName="Helvetica-Bold",
        alignment=TA_CENTER,
    )

    flow: list = []

    # ========= EN-TÊTE (5 cellules) =========
    logo_cell = ""
    if info.get("logo"):
        try:
            logo_cell = Image(info["logo"], width=22 * mm, height=16 * mm)
        except Exception:
            logo_cell = ""

    company_block = [
        Paragraph(info["nom"], st_company),
        Paragraph("Transport & Logistique — Conakry, Guinée", st_company_sub),
    ]

    # Cellule société = logo + texte en 2 sous-colonnes
    company_cell = Table(
        [[logo_cell, company_block]],
        colWidths=[24 * mm, None],
    )
    company_cell.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    header_row = [[
        company_cell,
        Paragraph("N° BON", st_hdr_label),
        Paragraph(f"<b>{bon.num_bon}</b>", st_hdr_val),
        Paragraph("DATE", st_hdr_label),
        Paragraph(bon.date.strftime("%d/%m/%Y"), st_hdr_val),
    ]]
    # Largeurs : société (≈44%) | label | valeur | label | valeur
    col_w_company = body_w * 0.44
    col_w_label = body_w * 0.08
    col_w_value = (body_w - col_w_company - 2 * col_w_label) / 2
    header_tbl = Table(
        header_row,
        colWidths=[col_w_company, col_w_label, col_w_value, col_w_label, col_w_value],
        rowHeights=[18 * mm],
    )
    header_tbl.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 1.0, NAVY),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, NAVY),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        # Bandes bleu foncé pour les labels
        ("BACKGROUND", (1, 0), (1, 0), NAVY),
        ("BACKGROUND", (3, 0), (3, 0), NAVY),
        # Zones valeur en bleu clair
        ("BACKGROUND", (2, 0), (2, 0), LIGHT),
        ("BACKGROUND", (4, 0), (4, 0), LIGHT),
        ("LEFTPADDING", (0, 0), (0, 0), 6),
        ("RIGHTPADDING", (0, 0), (0, 0), 6),
    ]))
    flow.append(header_tbl)

    # ========= BANDEAU TITRE =========
    title_tbl = Table(
        [[Paragraph("BON DE TRANSPORT BAUXITE  ——  FASTLANE LOGISTIC", st_title_band)]],
        colWidths=[body_w], rowHeights=[9 * mm],
    )
    title_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), NAVY),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOX", (0, 0), (-1, -1), 1.0, NAVY),
    ]))
    flow.append(title_tbl)
    flow.append(Spacer(1, 3 * mm))

    # ========= HELPERS =========
    def section_bar(text: str, width_mm: float) -> Table:
        t = Table(
            [[Paragraph(f"▌  {text}", st_section)]],
            colWidths=[width_mm], rowHeights=[6.5 * mm],
        )
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), NAVY),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ]))
        return t

    def kv_row(label: str, value, col_w_label: float, col_w_value: float) -> list:
        if value in (None, ""):
            value_txt = ""
        elif hasattr(value, "strftime"):
            value_txt = value.strftime("%H:%M")
        else:
            value_txt = str(value)
        return [
            Paragraph(label, st_label),
            Paragraph(value_txt, st_val),
        ]

    # Formatage d'une table label/valeur
    def build_kv_block(rows: list, width: float, col_ratio: float = 0.46) -> Table:
        col_w_l = width * col_ratio
        col_w_v = width - col_w_l
        data = [kv_row(lbl, val, col_w_l, col_w_v) for lbl, val in rows]
        t = Table(
            data, colWidths=[col_w_l, col_w_v],
            rowHeights=[7.5 * mm] * len(rows),
        )
        t.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.6, GREY_BORDER),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, GREY_BORDER),
            ("BACKGROUND", (0, 0), (0, -1), LIGHT),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]))
        return t

    # ========= BLOC PRINCIPAL : 2 COLONNES =========
    left_w = body_w * 0.42
    right_w = body_w - left_w - 3 * mm  # petit espace entre colonnes

    # ---- Colonne gauche : CHAUFFEUR / VÉHICULE / CLIENT-PROJET ----
    chauffeur_kv = build_kv_block([
        ("Prénom", bon.prenom),
        ("Nom", bon.nom),
        ("Téléphone", bon.telephone or ""),
    ], left_w)

    camion_code = bon.camion.code if bon.camion else ""
    plaque = bon.plaque or (bon.camion.plaque if bon.camion else "")
    vehicule_kv = build_kv_block([
        ("Plaque camion", plaque),
        ("Code camion", camion_code),
        ("Carte d'entrée", bon.carte_entree or ""),
        ("Lieu de chargement", bon.lieu_chargement or ""),
        ("Heure de départ", bon.heure_depart),
    ], left_w)

    projet_txt = ""
    if bon.contrat:
        projet_txt = getattr(bon.contrat, "code", None) or getattr(bon.contrat, "reference", "") or str(bon.contrat)
    client_kv = build_kv_block([
        ("Client / Projet", projet_txt),
    ], left_w)

    left_col_table = Table(
        [
            [section_bar("CHAUFFEUR", left_w)],
            [chauffeur_kv],
            [Spacer(1, 2 * mm)],
            [section_bar("VÉHICULE", left_w)],
            [vehicule_kv],
            [Spacer(1, 2 * mm)],
            [section_bar("CLIENT / PROJET", left_w)],
            [client_kv],
        ],
        colWidths=[left_w],
    )
    left_col_table.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    # ---- Colonne droite : PESÉES / OBSERVATIONS / SIGNATURES ----
    def _fmt_time(t):
        return t.strftime("%H:%M") if t else ""

    # Tableau 3 colonnes : PESÉE DÉBUT (T) | PESÉE FIN (T) | QUANTITÉ NETTE (T)
    quant_txt = f"{bon.quantite:.2f}" if bon.quantite else ""
    weigh_tbl = Table(
        [
            [
                Paragraph("PESÉE DÉBUT (T)", st_weigh_head),
                Paragraph("PESÉE FIN (T)", st_weigh_head),
                Paragraph("QUANTITÉ NETTE (T)", st_weigh_head),
            ],
            [
                Paragraph("", st_weigh_val),
                Paragraph("", st_weigh_val),
                Paragraph(quant_txt, st_weigh_val),
            ],
        ],
        colWidths=[right_w / 3] * 3,
        rowHeights=[6.5 * mm, 13 * mm],
    )
    weigh_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("BOX", (0, 0), (-1, -1), 0.6, GREY_BORDER),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, GREY_BORDER),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (2, 1), (2, 1), LIGHT),
    ]))

    # Horaires de pesée (infos système, si présentes)
    hr_start = _fmt_time(bon.heure_pesee_start)
    hr_end = _fmt_time(bon.heure_pesee_end)
    weigh_times = Table(
        [[
            Paragraph("HEURE PESÉE DÉBUT", st_weigh_head),
            Paragraph("HEURE PESÉE FIN", st_weigh_head),
        ],
        [
            Paragraph(hr_start, st_weigh_val),
            Paragraph(hr_end, st_weigh_val),
        ]],
        colWidths=[right_w / 2] * 2,
        rowHeights=[5.5 * mm, 9 * mm],
    )
    weigh_times.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3B6EA5")),
        ("BOX", (0, 0), (-1, -1), 0.6, GREY_BORDER),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, GREY_BORDER),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    # Zone OBSERVATIONS
    obs_txt = bon.observation or ""
    obs_box = Table(
        [[Paragraph(obs_txt or "&nbsp;", ParagraphStyle(
            "obs", parent=styles["Normal"], fontSize=9.5, leading=12,
            textColor=INK,
        ))]],
        colWidths=[right_w], rowHeights=[20 * mm],
    )
    obs_box.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.6, GREY_BORDER),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
    ]))

    # Zone SIGNATURES (3 colonnes : Chauffeur / Resp. site / Validation FASTLANE)
    sig_head = Table(
        [[
            Paragraph("Chauffeur", st_sig_role),
            Paragraph("Resp. site", st_sig_role),
            Paragraph("Validation FASTLANE", st_sig_role),
        ]],
        colWidths=[right_w / 3] * 3, rowHeights=[6 * mm],
    )
    sig_head.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), NAVY),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOX", (0, 0), (-1, -1), 0.6, NAVY),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.white),
    ]))

    sig_body = Table(
        [[
            Paragraph("Signature &amp; Date :", st_sig_note),
            Paragraph("Signature &amp; Cachet :", st_sig_note),
            Paragraph("Signature &amp; Cachet :", st_sig_note),
        ]],
        colWidths=[right_w / 3] * 3, rowHeights=[26 * mm],
    )
    sig_body.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.6, GREY_BORDER),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, GREY_BORDER),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
    ]))

    right_col_table = Table(
        [
            [section_bar("PESÉES ET QUANTITÉ TRANSPORTÉE", right_w)],
            [weigh_tbl],
            [Spacer(1, 1.5 * mm)],
            [weigh_times],
            [Spacer(1, 2 * mm)],
            [section_bar("OBSERVATIONS", right_w)],
            [obs_box],
            [Spacer(1, 2 * mm)],
            [section_bar("SIGNATURES ET VALIDATIONS", right_w)],
            [sig_head],
            [sig_body],
        ],
        colWidths=[right_w],
    )
    right_col_table.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    main_tbl = Table(
        [[left_col_table, "", right_col_table]],
        colWidths=[left_w, 3 * mm, right_w],
    )
    main_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    flow.append(main_tbl)
    flow.append(Spacer(1, 4 * mm))

    # ========= AVERTISSEMENT LÉGAL =========
    warn_tbl = Table(
        [[Paragraph(
            "⚠  Document officiel FASTLANE LOGISTIC — Toute altération est "
            "passible de poursuites. Conserver l'original.",
            st_warn,
        )]],
        colWidths=[body_w], rowHeights=[8 * mm],
    )
    warn_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FCE8E8")),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#C94B4B")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    flow.append(warn_tbl)

    # ========= FILIGRANE + PIED DE PAGE =========
    def _on_page(canvas, d):
        draw_watermark(canvas, pagesize)
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#888888"))
        footer = f"{info['nom']} — Bon de transport {bon.num_bon}"
        canvas.drawString(lm, 5 * mm, footer)
        canvas.drawRightString(
            page_w - rm, 5 * mm,
            f"Édité le {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        )
        canvas.restoreState()

    doc.build(flow, onFirstPage=_on_page, onLaterPages=_on_page)

    response = HttpResponse(buf.getvalue(), content_type="application/pdf")
    disposition = "inline" if inline else "attachment"
    response["Content-Disposition"] = (
        f'{disposition}; filename="bon_transport_{bon.num_bon}.pdf"'
    )
    return response
