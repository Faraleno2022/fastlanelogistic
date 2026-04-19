"""Utilitaires pour import Excel.

Chaque import a un **schéma** : liste de colonnes (label, clé, type, required).
Ce module fournit :
  - ``generate_import_template(schema, filename)`` — génère un XLSX vide avec
    les en-têtes attendues + une feuille d'instructions
  - ``read_excel_rows(uploaded_file, schema)`` — lit un fichier XLSX
    utilisateur et retourne ``(rows, errors)``
"""
from __future__ import annotations
import io
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from typing import Any, Callable, Sequence

from django.http import HttpResponse
from django.utils.text import slugify

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .exports import get_company_info


TYPE_PARSERS: dict[str, Callable[[Any], Any]] = {}


def parser(name: str):
    def deco(fn):
        TYPE_PARSERS[name] = fn
        return fn
    return deco


@parser("str")
def _p_str(v):
    if v is None:
        return ""
    return str(v).strip()


@parser("int")
def _p_int(v):
    if v is None or v == "":
        return None
    try:
        return int(str(v).strip().replace(" ", "").replace(",", ""))
    except (ValueError, TypeError):
        raise ValueError(f"entier attendu, reçu : {v!r}")


@parser("decimal")
def _p_decimal(v):
    if v is None or v == "":
        return Decimal(0)
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    try:
        cleaned = str(v).strip().replace(" ", "").replace(",", ".")
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        raise ValueError(f"nombre attendu, reçu : {v!r}")


@parser("date")
def _p_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"date attendue (jj/mm/aaaa), reçu : {v!r}")


@parser("time")
def _p_time(v):
    if v is None or v == "":
        return None
    if isinstance(v, time):
        return v
    if isinstance(v, datetime):
        return v.time()
    s = str(v).strip()
    for fmt in ("%H:%M", "%H:%M:%S", "%Hh%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    raise ValueError(f"heure attendue (HH:MM), reçu : {v!r}")


@parser("bool")
def _p_bool(v):
    if v is None or v == "":
        return False
    s = str(v).strip().lower()
    return s in ("1", "true", "vrai", "oui", "yes", "x", "o")


# ---------------------------------------------------------------------------
# Schéma
# ---------------------------------------------------------------------------

@dataclass
class Column:
    key: str            # nom du champ du modèle
    label: str          # en-tête Excel (ex: "Date prise (jj/mm/aaaa)")
    type: str = "str"   # str / int / decimal / date / time / bool
    required: bool = False
    help: str = ""
    # Optionnel : résolveur (ex: code camion -> instance Camion)
    resolve: Callable[[Any], Any] | None = None


# ---------------------------------------------------------------------------
# Génération du template
# ---------------------------------------------------------------------------

def generate_import_template(
    title: str, schema: Sequence[Column], *, filename: str | None = None,
    exemple_ligne: dict | None = None,
) -> HttpResponse:
    info = get_company_info()
    wb = Workbook()
    ws = wb.active
    ws.title = "Import"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    required_fill = PatternFill("solid", fgColor="F8CBAD")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14, color="1F4E78")
    hint_font = Font(italic=True, size=9, color="555555")
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # En-tête société
    ws["A1"] = info["nom"]
    ws["A1"].font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(3, len(schema)))
    ws["A2"] = f"Modèle d'import — {title}"
    ws["A2"].font = hint_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(3, len(schema)))
    ws["A3"] = ("Les colonnes en orange sont obligatoires. Ne renommez pas les en-têtes. "
                "Remplissez à partir de la ligne 6.")
    ws["A3"].font = hint_font
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max(3, len(schema)))

    # En-tête de colonnes (ligne 5)
    for i, col in enumerate(schema, start=1):
        c = ws.cell(row=5, column=i, value=col.label)
        c.fill = required_fill if col.required else header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = max(15, len(col.label) + 2)

    ws.row_dimensions[5].height = 32

    # Ligne d'exemple (facultative) en ligne 6
    if exemple_ligne:
        for i, col in enumerate(schema, start=1):
            ws.cell(row=6, column=i, value=exemple_ligne.get(col.key, ""))

    # Feuille d'instructions
    ws2 = wb.create_sheet("Instructions")
    ws2["A1"] = f"{info['nom']} — Instructions pour l'import : {title}"
    ws2["A1"].font = title_font
    ws2["A3"] = "Colonne"
    ws2["B3"] = "Obligatoire"
    ws2["C3"] = "Type attendu"
    ws2["D3"] = "Description"
    for cell in (ws2["A3"], ws2["B3"], ws2["C3"], ws2["D3"]):
        cell.fill = header_fill
        cell.font = header_font
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 60

    for i, col in enumerate(schema, start=4):
        ws2.cell(row=i, column=1, value=col.label)
        ws2.cell(row=i, column=2, value="OUI" if col.required else "non")
        ws2.cell(row=i, column=3, value=col.type)
        ws2.cell(row=i, column=4, value=col.help)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = (filename or f"modele_import_{slugify(title)}") + ".xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    return response


# ---------------------------------------------------------------------------
# Lecture d'un fichier utilisateur
# ---------------------------------------------------------------------------

def read_excel_rows(uploaded_file, schema: Sequence[Column]):
    """Lit le fichier utilisateur.

    Retourne ``(rows, errors)`` où :
      - rows    = liste de dicts {key: value}
      - errors  = liste de tuples (numero_ligne, message)
    """
    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb.active

    # Repérage de la ligne d'en-tête : on accepte ligne 1 OU ligne 5 (si template)
    header_row = None
    expected = {c.label.strip().lower() for c in schema}
    for r in range(1, 8):
        row_vals = [ws.cell(row=r, column=i).value for i in range(1, len(schema) + 1)]
        labels = {str(v).strip().lower() for v in row_vals if v}
        if expected.issubset(labels):
            header_row = r
            break
    if header_row is None:
        return [], [(0, "Impossible de trouver les en-têtes attendus dans les 7 premières lignes.")]

    # Mapping colonne index -> Column
    header_vals = [ws.cell(row=header_row, column=i).value for i in range(1, len(schema) + 10)]
    label_to_col = {c.label.strip().lower(): c for c in schema}
    idx_to_col: dict[int, Column] = {}
    for i, lbl in enumerate(header_vals, start=1):
        if lbl is None:
            continue
        col = label_to_col.get(str(lbl).strip().lower())
        if col:
            idx_to_col[i] = col

    missing = [c.label for c in schema if c.required and c.label.strip().lower() not in
               {str(v).strip().lower() for v in header_vals if v}]
    if missing:
        return [], [(header_row, f"Colonnes obligatoires manquantes : {', '.join(missing)}")]

    rows, errors = [], []
    first_data_row = header_row + 1
    # Si ligne 5 = header (template), ligne 6 peut être un exemple ; on lit depuis 6 dans tous les cas
    max_row = ws.max_row
    for r in range(first_data_row, max_row + 1):
        # ignore les lignes totalement vides
        if all(ws.cell(row=r, column=i).value in (None, "") for i in idx_to_col):
            continue
        row_data = {}
        row_ok = True
        for i, col in idx_to_col.items():
            raw = ws.cell(row=r, column=i).value
            if col.required and (raw is None or str(raw).strip() == ""):
                errors.append((r, f"{col.label} : valeur obligatoire manquante"))
                row_ok = False
                continue
            parser_fn = TYPE_PARSERS.get(col.type, _p_str)
            try:
                val = parser_fn(raw)
            except ValueError as e:
                errors.append((r, f"{col.label} : {e}"))
                row_ok = False
                continue
            # FK optionnelle : si valeur vide, on ne met rien dans row_data
            if col.resolve:
                if val in (None, ""):
                    continue  # laisse le champ à None/défaut du modèle
                try:
                    val = col.resolve(val)
                    if val is None:
                        errors.append((r, f"{col.label} : référence introuvable"))
                        row_ok = False
                        continue
                except Exception as e:
                    errors.append((r, f"{col.label} : {e}"))
                    row_ok = False
                    continue
            row_data[col.key] = val
        if row_ok:
            rows.append((r, row_data))
    return rows, errors
